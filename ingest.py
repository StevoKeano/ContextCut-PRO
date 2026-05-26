#!/usr/bin/env python3
"""
ContextCut — Qdrant ingest + file watcher

Embeds your markdown knowledge base into Qdrant using Voyage AI.

Usage:
  python ingest.py                        # one-shot ingest all .md files
  python ingest.py --watch                # ingest then watch for changes
  python ingest.py --query "your question" # test semantic search
  python ingest.py --clear                # wipe and recreate collection

Configuration via environment variables (all optional — defaults shown):
  CONTEXTCUT_QDRANT_HOST   localhost      Qdrant host
  CONTEXTCUT_QDRANT_PORT   6333           Qdrant port
  CONTEXTCUT_COLLECTION    contextcut     Qdrant collection name
  CONTEXTCUT_KB_DIR        ~/contextcut/knowledge  Directory to watch
  VOYAGE_API_KEY           (required)     Voyage AI API key

Notes:
  - Only .md files are ingested
  - Files matching EXCLUDE_FILES or containing .bak- are skipped
  - Voyage AI free tier: ~3 RPM — a 21s delay is added between embeds
  - First 4000 chars of each file are stored as payload text
  - First 8000 chars are embedded (Voyage model input limit)
"""

import os
import sys
import time
import json
import random
import hashlib
import urllib.request
import argparse
from pathlib import Path

VOYAGE_AVAILABLE = False
try:
    import voyageai
    VOYAGE_AVAILABLE = True
except ImportError:
    pass

from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct, PointIdsList, Filter, FieldCondition, MatchValue

# ── Config ────────────────────────────────────────────────────────────────────
QDRANT_HOST  = os.getenv("CONTEXTCUT_QDRANT_HOST", "localhost")
QDRANT_PORT  = int(os.getenv("CONTEXTCUT_QDRANT_PORT", "6333"))
COLLECTION   = os.getenv("CONTEXTCUT_COLLECTION",  "contextcut")
KB_DIR       = Path(os.getenv("CONTEXTCUT_KB_DIR", str(Path.home() / "contextcut" / "knowledge"))).expanduser()
VOYAGE_API_KEY = os.environ.get("VOYAGE_API_KEY", "").strip().strip('"').strip("'")
OLLAMA_EMBED   = os.environ.get("CONTEXTCUT_EMBED_MODEL", "").strip().strip('"').strip("'")
EMBED_MODE     = os.environ.get("CONTEXTCUT_EMBED_MODE", "").strip().strip('"').strip("'")
OLLAMA_URL     = os.environ.get("CONTEXTCUT_UPSTREAM", "http://localhost:11434")
VOYAGE_MODEL = "voyage-3"

def _get_embed_dim():
    """Return embedding dimension based on active mode and model."""
    if EMBED_MODE == "voyage" and VOYAGE_API_KEY:
        return 1024
    if OLLAMA_EMBED:
        dims = {
            "nomic-embed-text": 768,
            "nomic-embed-text-v1.5": 768,
            "nomic-embed-text-v2-moe": 768,
            "mxbai-embed-large": 1024,
            "bge-m3": 1024,
            "qwen3-embedding:0.6b": 4096,
            "qwen3-embedding:4b": 4096,
            "qwen3-embedding:8b": 4096,
            "snowflake-arctic-embed-l": 1024,
            "all-minilm": 384,
        }
        base = OLLAMA_EMBED.split(":")[0]
        return dims.get(OLLAMA_EMBED, dims.get(base, 1024))
    return 1024

EMBED_DIM = _get_embed_dim()

# Per-model max input tokens for embedding (~25% below actual limit to account for tokenizer mismatch)
def _get_embed_max_tokens():
    if EMBED_MODE == "voyage" and VOYAGE_API_KEY:
        return 30000
    if OLLAMA_EMBED:
        ctx = {
            "nomic-embed-text": 1500,
            "nomic-embed-text-v1.5": 6000,
            "nomic-embed-text-v2-moe": 6000,
            "mxbai-embed-large": 400,
            "bge-m3": 6000,
            "qwen3-embedding:0.6b": 24000,
            "qwen3-embedding:4b": 24000,
            "qwen3-embedding:8b": 24000,
            "snowflake-arctic-embed-l": 6000,
            "all-minilm": 6000,
        }
        base = OLLAMA_EMBED.split(":")[0]
        return ctx.get(OLLAMA_EMBED, ctx.get(base, 1500))
    return 1500

MAX_EMBED_TOKENS = _get_embed_max_tokens()

CHUNK_TOKENS = 512
CHUNK_OVERLAP = 50

def chunk_text(text: str) -> list[str]:
    try:
        import tiktoken
        enc = tiktoken.get_encoding("cl100k_base")
        tokens = enc.encode(text)
    except ImportError:
        return [text[:MAX_EMBED_TOKENS]]
    if len(tokens) <= CHUNK_TOKENS:
        return [text]
    chunks = []
    i = 0
    while i < len(tokens):
        chunk_tokens = tokens[i:i + CHUNK_TOKENS]
        chunks.append(enc.decode(chunk_tokens))
        i += CHUNK_TOKENS - CHUNK_OVERLAP
    return chunks

# Files to never ingest
EXCLUDE_FILES = {"MEMORY.md", "MEMORY.txt", "MEMORY.py"}

# Allowed file extensions for knowledge base ingestion
ALLOWED_EXT = {
    ".md", ".txt", ".py", ".js", ".ts", ".html", ".css",
    ".csv", ".json", ".xml", ".yaml", ".yml",
    ".go", ".rs", ".rb", ".java", ".c", ".cpp", ".h",
    ".sh", ".sql", ".log",
    ".pdf", ".docx", ".xlsx",
}

last_embed_time = 0

# ── Clients ───────────────────────────────────────────────────────────────────
vc = voyageai.Client(api_key=VOYAGE_API_KEY) if (VOYAGE_API_KEY and VOYAGE_AVAILABLE) else None
qc = QdrantClient(host=QDRANT_HOST, port=QDRANT_PORT)

def _ollama_embed(texts, model):
    """Embed using Ollama's /api/embed endpoint."""
    payloads = []
    for t in texts:
        try:
            payload = json.dumps({"model": model, "input": t}).encode()
        except ValueError as e:
            print(f"  [!] Ollama embed JSON encoding error: {e}")
            print(f"  [!] Input text contains invalid characters for JSON")
            print(f"  [!] Repr of first 200 chars: {repr(t[:200])}")
            raise
        req = urllib.request.Request(f"{OLLAMA_URL}/api/embed", data=payload, method="POST")
        req.add_header("Content-Type", "application/json")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            print(f"  [!] Ollama HTTP {e.code}: {body}")
            raise
        embeddings = data.get("embeddings", [])
        if embeddings:
            payloads.append(embeddings[0])
        else:
            payloads.append([0.0] * EMBED_DIM)
    return payloads

# ── Collection ────────────────────────────────────────────────────────────────
def ensure_collection():
    existing = [c.name for c in qc.get_collections().collections]
    if COLLECTION not in existing:
        try:
            qc.create_collection(
                collection_name=COLLECTION,
                vectors_config=VectorParams(size=EMBED_DIM, distance=Distance.COSINE),
            )
            print(f"[+] Created collection '{COLLECTION}' (dim={EMBED_DIM})")
        except Exception:
            # Already exists (race with another process)
            pass
    else:
        info = qc.get_collection(COLLECTION)
        actual_dim = info.config.params.vectors.size
        if actual_dim != EMBED_DIM:
            print(f"[!] Dimension mismatch: Qdrant has {actual_dim}, model needs {EMBED_DIM}")
            print(f"[!] Fix: update CONTEXTCUT_EMBED_MODE/CONTEXTCUT_EMBED_MODEL in .env or use the dashboard Settings")
            raise SystemExit(1)

# ── Ingest ────────────────────────────────────────────────────────────────────
def file_id(path: Path) -> str:
    return hashlib.md5(str(path).encode()).hexdigest()

def sanitize_text(text: str) -> str:
    return text.replace('\r\n', '\n').strip()

processing_queue = set()

def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(path)
            return "\n".join(page.get_text() for page in doc)
        except (ImportError, Exception) as e:
            print(f"  [!] PDF extraction failed for {path.name}: {e}")
            return ""
    elif ext == ".docx":
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        except (ImportError, Exception) as e:
            print(f"  [!] DOCX extraction failed for {path.name}: {e}")
            return ""
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            rows = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(rows)
        except (ImportError, Exception) as e:
            print(f"  [!] XLSX extraction failed for {path.name}: {e}")
            return ""
    else:
        return path.read_text(encoding="utf-8", errors="ignore").strip()

def ingest_file(path: Path):
    if path.name in processing_queue:
        return
    processing_queue.add(path.name)
    
    try:
        if not path.exists:
            return
        raw_text = extract_text(path)
        if not raw_text:
            return
        
        chunks = chunk_text(raw_text)
        clean_chunks = [c for c in (sanitize_text(c) for c in chunks) if c]
        
        if not clean_chunks:
            print(f"  [!] {path.name}: all chunks empty after sanitization, skipping")
            return
        
        result = safe_embed(clean_chunks, model=VOYAGE_MODEL, input_type="document")
        
        if len(result.embeddings) != len(clean_chunks):
            print(f"  [!] {path.name}: expected {len(clean_chunks)} embeddings, got {len(result.embeddings)}")
            n = min(len(result.embeddings), len(clean_chunks))
            clean_chunks = clean_chunks[:n]
            result.embeddings = result.embeddings[:n]
        
        fid = file_id(path)
        points = []
        for i, chunk_text_str in enumerate(clean_chunks):
            chunk_id = int(hashlib.md5((fid + str(i)).encode()).hexdigest()[:8], 16)
            points.append(PointStruct(
                id=chunk_id,
                vector=result.embeddings[i],
                payload={
                    "filename": path.name,
                    "path": str(path),
                    "chunk_index": i,
                    "total_chunks": len(clean_chunks),
                    "text": chunk_text_str[:4000],
                }
            ))
        
        qc.upsert(collection_name=COLLECTION, points=points)
        info = f"{len(points)} chunk(s)" if len(points) > 1 else "1 chunk"
        print(f"  [ok] {path.name} ({info})")
    finally:
        processing_queue.remove(path.name)


def safe_embed(texts, model, input_type):
    global last_embed_time

    # Use Ollama embed if mode is ollama or if Voyage is unavailable
    if (EMBED_MODE == "ollama" and OLLAMA_EMBED) or (not VOYAGE_API_KEY or not VOYAGE_AVAILABLE):
        if OLLAMA_EMBED:
            try:
                vectors = _ollama_embed(texts, OLLAMA_EMBED)
                class EmbedResult:
                    pass
                result = EmbedResult()
                result.embeddings = vectors
                return result
            except Exception as e:
                print(f" [!] Ollama embed error: {e}")
                raise

    elapsed = time.time() - last_embed_time
    if elapsed < 22:
        time.sleep(22 - elapsed)

    try:
        result = vc.embed(texts, model=model, input_type=input_type)
        last_embed_time = time.time()
        return result
    except voyageai.error.RateLimitError:
        wait_time = 60 + random.uniform(5, 15)
        print(f" [!] Rate limit hit, backing off for {wait_time:.1f}s...")
        time.sleep(wait_time)
        return safe_embed(texts, model, input_type)

def should_ingest(path: Path) -> bool:
    return (
        path.suffix.lower() in ALLOWED_EXT
        and path.name not in EXCLUDE_FILES
        and ".bak-" not in path.name
    )

def ingest_all():
    ensure_collection()
    md_files = [f for f in KB_DIR.iterdir() if f.is_file() and should_ingest(f)]
    if not md_files:
        print(f"No eligible files found in {KB_DIR}")
        return
    print(f"[*] Ingesting {len(md_files)} file(s) from {KB_DIR} ...")
    for f in md_files:
        ingest_file(f)
    print("[*] Done.")

# ── Query ─────────────────────────────────────────────────────────────────────
def query(q: str, top_k: int = 5):
    ensure_collection()
    result = vc.embed([q], model=VOYAGE_MODEL, input_type="query")
    vector = result.embeddings[0]
    hits   = qc.query_points(collection_name=COLLECTION, query=vector, limit=top_k).points
    print(f"\nTop {top_k} results for: \"{q}\"\n")
    for i, hit in enumerate(hits, 1):
        print(f"  {i}. [{hit.score:.3f}] {hit.payload.get('filename','?')}")
        print(f"     {hit.payload.get('text','')[:200].strip()}\n")

# ── Clear ─────────────────────────────────────────────────────────────────────
def clear():
    qc.delete_collection(COLLECTION)
    print(f"[+] Collection '{COLLECTION}' deleted.")

# ── Watch ─────────────────────────────────────────────────────────────────────
def watch():
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except ImportError:
        print("ERROR: watchdog not installed. Run: pip install watchdog")
        sys.exit(1)

    class Handler(FileSystemEventHandler):
        def __init__(self):
            self._last = {}

        def _debounce(self, path, secs=30) -> bool:
            now = time.time()
            if now - self._last.get(path, 0) < secs:
                return False
            self._last[path] = now
            return True

        def on_modified(self, event):
            p = Path(event.src_path)
            if should_ingest(p) and self._debounce(event.src_path):
                ingest_file(p)

        def on_created(self, event):
            p = Path(event.src_path)
            if should_ingest(p) and self._debounce(event.src_path):
                ingest_file(p)

        def on_deleted(self, event):
            p = Path(event.src_path)
            if should_ingest(p) and self._debounce(event.src_path):
                try:
                    qc.delete(
                        collection_name=COLLECTION,
                        points_selector=Filter(
                            must=[FieldCondition(key="filename", match=MatchValue(value=p.name))]
                        )
                    )
                    print(f"  [del] {p.name} — removed from Qdrant (all chunks)")
                except Exception as e:
                    print(f"  [!] Failed to remove {p.name} from Qdrant: {e}")

    ensure_collection()
    ingest_all()
    observer = Observer()
    observer.schedule(Handler(), str(KB_DIR), recursive=False)
    observer.start()
    print(f"[*] Watching {KB_DIR} for changes. Ctrl+C to stop.")
    try:
        while True:
            time.sleep(2)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="ContextCut ingest tool")
    parser.add_argument("--watch", action="store_true", help="Watch KB_DIR for changes")
    parser.add_argument("--query", type=str,            help="Test semantic search")
    parser.add_argument("--clear", action="store_true", help="Wipe Qdrant collection")
    args = parser.parse_args()

    if args.clear:
        clear()
        sys.exit(0)
    if args.query:
        if not (EMBED_MODE == "ollama" and OLLAMA_EMBED) and not (VOYAGE_API_KEY and VOYAGE_AVAILABLE):
            print("ERROR: No embedding backend configured.")
            print("  Set VOYAGE_API_KEY or CONTEXTCUT_EMBED_MODEL")
            sys.exit(1)
        query(args.query)
        sys.exit(0)

    # watch or ingest_all need embed backend
    if EMBED_MODE == "voyage" and VOYAGE_API_KEY:
        print(f"[ingest] Embedding: Voyage AI (voyage-3)")
    elif EMBED_MODE == "ollama" and OLLAMA_EMBED:
        print(f"[ingest] Embedding: Ollama ({OLLAMA_EMBED})")
    elif VOYAGE_API_KEY:
        print(f"[ingest] Embedding: Voyage AI (voyage-3)")
    elif OLLAMA_EMBED:
        print(f"[ingest] Embedding: Ollama ({OLLAMA_EMBED})")
    else:
        print("ERROR: Neither VOYAGE_API_KEY nor CONTEXTCUT_EMBED_MODEL set.")
        print("  export VOYAGE_API_KEY=your-key-here")
        print("  export CONTEXTCUT_EMBED_MODEL=nomic-embed-text")
        sys.exit(1)

    if not KB_DIR.exists():
        print(f"ERROR: Knowledge base directory not found: {KB_DIR}")
        print(f"  Create it: mkdir -p {KB_DIR}")
        print(f"  Or set: export CONTEXTCUT_KB_DIR=/path/to/your/markdown/files")
        sys.exit(1)

    if args.watch:
        watch()
    else:
        ingest_all()
