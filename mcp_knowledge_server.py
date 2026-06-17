#!/usr/bin/env python3
"""
MCP Knowledge Server for ContextCut-PRO.

Exposes knowledge base search and file ingestion via Model Context Protocol,
usable by Claude Desktop, Cursor, VS Code, ChatGPT, Gemini, and any MCP client.

Usage:
    python mcp_knowledge_server.py                        # stdio transport
    python mcp_knowledge_server.py --transport http       # Streamable HTTP
    python mcp_knowledge_server.py --transport http --port 8910

Environment variables (inherited from ContextCut-PRO):
    CONTEXTCUT_QDRANT_HOST, CONTEXTCUT_QDRANT_PORT, CONTEXTCUT_COLLECTION,
    CONTEXTCUT_KB_DIR, CONTEXTCUT_EMBED_MODE, CONTEXTCUT_EMBED_MODEL,
    CONTEXTCUT_UPSTREAM, VOYAGE_API_KEY
"""

import hashlib
import json
import os
import sys
import time
import urllib.request
from pathlib import Path

try:
    from fastmcp import FastMCP
except ImportError:
    print("ERROR: fastmcp not installed. Run: pip install fastmcp")
    sys.exit(1)

from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct, Filter, FieldCondition, MatchValue

from qdrant_proxy_final import (
    qdrant_context,
    make_stats_json as _proxy_stats,
    KB_DIR,
    QDRANT_HOST,
    QDRANT_PORT,
    COLLECTION,
    ALLOWED_EXT,
    MIN_SCORE,
    TOP_K,
    UPSTREAM,
    _EMBED_MODE,
    _LOCAL_EMBED,
    _VK,
    _VOYAGE_AVAILABLE,
)

mcp = FastMCP("contextcut-knowledge")

_qclient = None
_vc = None
_last_embed_ts = 0.0


def _get_qclient():
    global _qclient
    if _qclient is None:
        _qclient = QdrantClient(host=QDRANT_HOST, port=QDRANT_PORT)
    return _qclient


EXCLUDE_FILES = {"MEMORY.md", "MEMORY.txt", "MEMORY.py"}


def _should_ingest(path: Path) -> bool:
    return (
        path.suffix.lower() in ALLOWED_EXT
        and path.name not in EXCLUDE_FILES
        and ".bak-" not in path.name
    )


def _file_id(path: Path) -> str:
    return hashlib.md5(str(path).encode()).hexdigest()


def _file_content_hash(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def _sanitize_text(text: str) -> str:
    return text.replace('\r\n', '\n').strip()


def _is_file_current(path: Path) -> bool:
    qc = _get_qclient()
    ch = _file_content_hash(path)
    result, _ = qc.scroll(
        collection_name=COLLECTION,
        scroll_filter=Filter(
            must=[
                FieldCondition(key="filename", match=MatchValue(value=path.name)),
                FieldCondition(key="content_hash", match=MatchValue(value=ch)),
            ]
        ),
        limit=1,
    )
    return len(result) > 0


def _extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        try:
            import fitz
        except ImportError:
            return ""
        try:
            doc = fitz.open(path)
            return "\n".join(page.get_text() for page in doc)
        except Exception:
            return ""
    elif ext == ".docx":
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            return ""
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            rows = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(rows)
        except ImportError:
            return ""
    else:
        return path.read_text(encoding="utf-8", errors="ignore").strip()


def _chunk_text(text: str, max_tokens: int = 512, overlap: int = 50) -> list[str]:
    try:
        import tiktoken
        enc = tiktoken.get_encoding("cl100k_base")
        tokens = enc.encode(text)
    except ImportError:
        return [text[:8000]]
    if len(tokens) <= max_tokens:
        return [text]
    chunks = []
    i = 0
    while i < len(tokens):
        chunk_tokens = tokens[i:i + max_tokens]
        chunks.append(enc.decode(chunk_tokens))
        i += max_tokens - overlap
    return chunks


def _batch_embed(texts: list[str], input_type: str = "document") -> list[list[float]]:
    """Embed a list of texts using the configured backend. Returns list of vectors."""
    global _vc, _last_embed_ts

    use_voyage = _EMBED_MODE == "voyage" and _VK and _VOYAGE_AVAILABLE
    use_ollama = bool(_LOCAL_EMBED)

    if use_voyage:
        if _vc is None:
            import voyageai
            _vc = voyageai.Client(api_key=_VK)
        elapsed = time.time() - _last_embed_ts
        if elapsed < 22:
            time.sleep(22 - elapsed)
        try:
            result = _vc.embed(texts, model="voyage-3", input_type=input_type)
            _last_embed_ts = time.time()
            return result.embeddings
        except Exception as e:
            err_msg = str(e).lower()
            if "rate" in err_msg or "429" in err_msg:
                time.sleep(60)
                return _batch_embed(texts, input_type)
            if use_ollama:
                print(f"[mcp] Voyage failed, falling back to Ollama: {e}")
            else:
                raise

    if use_ollama:
        vectors = []
        for t in texts:
            try:
                payload = json.dumps({"model": _LOCAL_EMBED, "input": t}).encode()
                req = urllib.request.Request(
                    f"{UPSTREAM}/api/embed", data=payload, method="POST"
                )
                req.add_header("Content-Type", "application/json")
                with urllib.request.urlopen(req, timeout=60) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                embeddings = data.get("embeddings", [])
                if embeddings:
                    vectors.append(embeddings[0])
                else:
                    vectors.append([0.0] * 1024)
            except Exception as e:
                print(f"[mcp] Ollama embed error: {e}")
                vectors.append([0.0] * 1024)
        return vectors

    raise RuntimeError("No embedding backend configured (set VOYAGE_API_KEY or CONTEXTCUT_EMBED_MODEL)")


def _ensure_collection():
    qc = _get_qclient()
    existing = [c.name for c in qc.get_collections().collections]
    if COLLECTION not in existing:
        qc.create_collection(
            collection_name=COLLECTION,
            vectors_config=VectorParams(size=1024, distance=Distance.COSINE),
        )


# ── Tools ─────────────────────────────────────────────────────────────────


@mcp.tool()
def knowledge_search(query: str, top_k: int = 5) -> str:
    """Search the knowledge base for context relevant to a query. Returns formatted text with sources and scores."""
    ctx, meta = qdrant_context(query)
    if not ctx:
        return "No relevant context found."
    parts = []
    for m in meta:
        parts.append(f"[score={m['score']:.3f}] source={m['source']}")
    return "\n".join(parts) + "\n\n---\n\n" + ctx


@mcp.tool()
def knowledge_search_structured(query: str, top_k: int = 5) -> list:
    """Search the knowledge base and return structured results as a list of {source, score, text} objects."""
    ctx, meta = qdrant_context(query)
    if not meta:
        return []
    texts = ctx.split("\n\n---\n\n")
    results = []
    for i, m in enumerate(meta):
        results.append({
            "source": m["source"],
            "score": m["score"],
            "text": texts[i] if i < len(texts) else "",
        })
    return results


@mcp.tool()
def knowledge_list_files() -> list:
    """List all eligible knowledge files in the KB directory."""
    if not KB_DIR.exists():
        return []
    files = []
    for f in sorted(KB_DIR.iterdir()):
        if f.is_file() and _should_ingest(f):
            stat = f.stat()
            files.append({
                "name": f.name,
                "path": str(f),
                "size": stat.st_size,
                "modified": stat.st_mtime,
            })
    return files


@mcp.tool()
def knowledge_stats() -> dict:
    """Return KB statistics: total chunks, file count, proxy stats, embed mode."""
    qc = _get_qclient()
    try:
        collection_info = qc.get_collection(COLLECTION)
        total_chunks = collection_info.points_count
    except Exception:
        total_chunks = 0
    files = knowledge_list_files()
    proxy = _proxy_stats()
    return {
        "total_chunks": total_chunks,
        "total_files": len(files),
        "embed_mode": os.environ.get("CONTEXTCUT_EMBED_MODE", "voyage"),
        "collection": COLLECTION,
        "kb_dir": str(KB_DIR),
        "qdrant": f"{QDRANT_HOST}:{QDRANT_PORT}",
        "proxy_requests": proxy.get("total_requests", 0),
        "proxy_tokens_saved": proxy.get("total_saved", 0),
    }


@mcp.tool()
def ingest_file(filename: str) -> str:
    """Ingest a single file (by name, relative to KB_DIR) into the Qdrant vector store. Re-embeds all chunks."""
    path = (KB_DIR / filename).resolve()
    kb_resolved = KB_DIR.resolve()
    if not str(path).startswith(str(kb_resolved)):
        return f"ERROR: path must be under {KB_DIR}"
    if not path.exists():
        return f"ERROR: file not found: {path}"
    if not _should_ingest(path):
        return f"ERROR: {path.suffix} files are not eligible for ingestion"

    raw_text = _extract_text(path)
    if not raw_text:
        return f"Skipped {path.name}: no text extracted"
    chunks = _chunk_text(raw_text)
    clean_chunks = [c for c in (_sanitize_text(c) for c in chunks) if c]
    if not clean_chunks:
        return f"Skipped {path.name}: all chunks empty"

    embeddings = _batch_embed(clean_chunks)
    n = min(len(embeddings), len(clean_chunks))
    clean_chunks = clean_chunks[:n]
    embeddings = embeddings[:n]

    qc = _get_qclient()
    fid = _file_id(path)
    ch = _file_content_hash(path)
    points = []
    for i, text_str in enumerate(clean_chunks):
        chunk_id = int(hashlib.md5((fid + str(i)).encode()).hexdigest()[:8], 16)
        points.append(PointStruct(
            id=chunk_id,
            vector=embeddings[i],
            payload={
                "filename": path.name,
                "path": str(path),
                "chunk_index": i,
                "total_chunks": len(clean_chunks),
                "content_hash": ch,
                "text": text_str[:4000],
            },
        ))
    qc.upsert(collection_name=COLLECTION, points=points)
    return f"Ingested {path.name}: {len(points)} chunk(s)"


@mcp.tool()
def ingest_all() -> str:
    """Ingest all changed or new files in KB_DIR into Qdrant."""
    if not KB_DIR.exists():
        return f"ERROR: KB_DIR not found: {KB_DIR}"

    _ensure_collection()
    files = [f for f in sorted(KB_DIR.iterdir()) if f.is_file() and _should_ingest(f)]
    if not files:
        return f"No eligible files in {KB_DIR}"

    results = []
    for f in files:
        if _is_file_current(f):
            results.append(f"{f.name}: unchanged, skipped")
            continue
        result = ingest_file(f.name)
        results.append(result)
    return "\n".join(results)


@mcp.tool()
def ingest_status() -> list:
    """Show ingest status for each file: current (up-to-date) or stale (needs re-ingest)."""
    if not KB_DIR.exists():
        return []
    statuses = []
    for f in sorted(KB_DIR.iterdir()):
        if f.is_file() and _should_ingest(f):
            statuses.append({
                "name": f.name,
                "current": _is_file_current(f),
            })
    return statuses


# ── Resources ─────────────────────────────────────────────────────────────


@mcp.resource("knowledge://files")
def resource_files() -> str:
    return json.dumps(knowledge_list_files(), indent=2)


@mcp.resource("knowledge://stats")
def resource_stats() -> str:
    return json.dumps(knowledge_stats(), indent=2)


@mcp.resource("knowledge://search/{query}")
def resource_search(query: str) -> str:
    return json.dumps(knowledge_search_structured(query), indent=2)


# ── Main ──────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="ContextCut-PRO MCP Knowledge Server")
    parser.add_argument("--transport", choices=["stdio", "http"], default="stdio")
    parser.add_argument("--port", type=int, default=8910)
    args = parser.parse_args()
    mcp.run(transport=args.transport, port=args.port)
